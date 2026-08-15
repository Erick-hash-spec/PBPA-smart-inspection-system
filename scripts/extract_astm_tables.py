import openpyxl, json, sys

wb = openpyxl.load_workbook('SHORE TANK CALCULATION EXCELL.xlsx', data_only=True)

def extract_table(sheet_name):
    ws = wb[sheet_name]
    # Row 9 = temperature headers (col index 3 onwards, i.e. col D=index 3)
    temp_row = list(ws.iter_rows(min_row=9, max_row=9, values_only=True))[0]
    # temperatures start at column index 3 (0-based), which is col D
    temps = []
    for v in temp_row[3:]:
        if v is not None:
            temps.append(float(v))

    # Data rows start at row 10
    table = {}  # density_key -> {temp -> value}
    for row in ws.iter_rows(min_row=10, max_row=ws.max_row, values_only=True):
        density_key = row[2]  # col C = density key
        if density_key is None or not isinstance(density_key, (int, float)):
            continue
        density_key = round(float(density_key), 4)
        values = row[3:]  # col D onwards
        row_dict = {}
        for i, v in enumerate(values):
            if i < len(temps) and v is not None:
                row_dict[temps[i]] = round(float(v), 6)
        if row_dict:
            table[density_key] = row_dict
    return temps, table

temps59, table59 = extract_table('Table 59B')
temps60, table60 = extract_table('Table 60B')

print(f"Table 59B: {len(table59)} density rows, {len(temps59)} temperature cols")
print(f"  Density range: {min(table59.keys()):.4f} - {max(table59.keys()):.4f}")
print(f"  Temp range: {min(temps59)} - {max(temps59)}")
print(f"Table 60B: {len(table60)} density rows, {len(temps60)} temperature cols")
print(f"  Density range: {min(table60.keys()):.4f} - {max(table60.keys()):.4f}")
print(f"  Temp range: {min(temps60)} - {max(temps60)}")

# Verify against sample document
# Sample: density=0.735, sample_temp=30 -> d20 expected=0.7439
d_key = 0.735
t_key = 30.0
if d_key in table59 and t_key in table59[d_key]:
    print(f"\nTable 59B lookup: density=0.735, temp=30 -> {table59[d_key][t_key]} (expected 0.7439)")
else:
    # Find nearest
    keys = sorted(table59.keys())
    nearest = min(keys, key=lambda k: abs(k - d_key))
    print(f"\nNearest density key to 0.735: {nearest}")
    print(f"  Available temps near 30: {[t for t in sorted(table59[nearest].keys()) if 28 <= t <= 32]}")

# Verify VCF: d20=0.7439, tank_temp=29 -> VCF expected=0.9890
d20_key = 0.7439
t_key2 = 29.0
if d20_key in table60 and t_key2 in table60[d20_key]:
    print(f"Table 60B lookup: d20=0.7439, tank_temp=29 -> {table60[d20_key][t_key2]} (expected 0.9890)")
else:
    keys60 = sorted(table60.keys())
    nearest60 = min(keys60, key=lambda k: abs(k - d20_key))
    print(f"\nNearest d20 key to 0.7439: {nearest60}")
    if t_key2 in table60[nearest60]:
        print(f"  VCF at temp=29: {table60[nearest60][t_key2]}")

# Save as JSON for inspection
with open('astm_table59b.json', 'w') as f:
    json.dump({str(k): v for k, v in table59.items()}, f)
with open('astm_table60b.json', 'w') as f:
    json.dump({str(k): v for k, v in table60.items()}, f)
print("\nSaved astm_table59b.json and astm_table60b.json")
