import openpyxl
from openpyxl.styles import colors

# Load the workbook
wb = openpyxl.load_workbook('SHORE TANK CALCULATION EXCELL.xlsx')
print(f'Sheet names: {wb.sheetnames}')

# Get the first sheet
ws = wb.active
print(f'\nActive sheet: {ws.title}')
print('\n' + '='*80)
print('CELL VALUES AND COLORS')
print('='*80 + '\n')

# Extract cells with values and their colors (limited rows for clarity)
for row in ws.iter_rows(min_row=1, max_row=100, min_col=1, max_col=10):
    for cell in row:
        if cell.value:
            try:
                fill = cell.fill
                if fill and fill.fgColor:
                    color = fill.fgColor.rgb
                else:
                    color = 'None'
            except:
                color = 'Unknown'
            
            print(f'{cell.coordinate}: {cell.value} | Color: {color} | Type: {cell.data_type}')
            
            # If it's a formula, show it
            if cell.data_type == 'f':
                print(f'  -> Formula: {cell.value}')

print('\n' + '='*80)
print('FORMULAS')
print('='*80 + '\n')

# Extract only formulas
for row in ws.iter_rows(min_row=1, max_row=100, min_col=1, max_col=10):
    for cell in row:
        if cell.data_type == 'f':
            print(f'{cell.coordinate}: {cell.value}')
