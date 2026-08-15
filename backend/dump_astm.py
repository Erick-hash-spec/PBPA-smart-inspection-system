import openpyxl

wb = openpyxl.load_workbook('SHORE TANK CALCULATION EXCELL.xlsx', data_only=True)

for sheet_name in ['Table 59B', 'Table 60B', 'ASTM Calculator']:
    ws = wb[sheet_name]
    print(f"\n{'='*60}")
    print(f"SHEET: {sheet_name}  (rows={ws.max_row}, cols={ws.max_column})")
    print(f"{'='*60}")
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 30), values_only=True)):
        if any(v is not None for v in row):
            print(f"Row {i+1}: {row}")
