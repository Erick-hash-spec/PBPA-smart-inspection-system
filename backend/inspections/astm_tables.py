"""
ASTM D1250 / API MPMS Ch.11 Table Lookup Engine.

Table 59B: observed density (kg/L) + sample temperature (C) -> density @20C
Table 60B: density @20C (kg/L) + tank temperature (C) -> VCF
"""

import bisect
import json
import os

_TABLE_59B = None
_TABLE_60B = None
_TABLE_59B_KEYS = []
_TABLE_60B_KEYS = []
_TABLE_59B_TEMP_KEYS = {}
_TABLE_60B_TEMP_KEYS = {}
_TABLE_59B_RANGE = None
_TABLE_60B_RANGE = None

_BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
_EXCEL_PATH = os.path.join(_BASE_DIR, 'SHORE TANK CALCULATION EXCELL.xlsx')
_JSON_59B_PATH = os.path.join(_BASE_DIR, 'astm_table59b.json')
_JSON_60B_PATH = os.path.join(_BASE_DIR, 'astm_table60b.json')


def _load_tables():
    if _TABLE_59B is not None:
        return

    try:
        _set_table_cache(_load_json_table(_JSON_59B_PATH), _load_json_table(_JSON_60B_PATH))
        return
    except Exception:
        pass

    try:
        import openpyxl

        wb = openpyxl.load_workbook(_EXCEL_PATH, data_only=True, read_only=True)
        try:
            _set_table_cache(_parse_sheet(wb['Table 59B']), _parse_sheet(wb['Table 60B']))
        finally:
            wb.close()
    except Exception:
        # Tables unavailable; callers can use formula fallback.
        _set_table_cache({}, {})


def _load_json_table(path):
    with open(path, 'r', encoding='utf-8') as f:
        raw = json.load(f)

    return {
        round(float(density), 4): {float(temp): float(value) for temp, value in row.items()}
        for density, row in raw.items()
    }


def _set_table_cache(table59, table60):
    global _TABLE_59B, _TABLE_60B
    global _TABLE_59B_KEYS, _TABLE_60B_KEYS
    global _TABLE_59B_TEMP_KEYS, _TABLE_60B_TEMP_KEYS
    global _TABLE_59B_RANGE, _TABLE_60B_RANGE

    _TABLE_59B = table59
    _TABLE_60B = table60
    _TABLE_59B_KEYS, _TABLE_59B_TEMP_KEYS, _TABLE_59B_RANGE = _build_meta(table59)
    _TABLE_60B_KEYS, _TABLE_60B_TEMP_KEYS, _TABLE_60B_RANGE = _build_meta(table60)


def _build_meta(table):
    if not table:
        return [], {}, None

    density_keys = sorted(table.keys())
    temp_keys_by_density = {density: sorted(row.keys()) for density, row in table.items()}
    all_temps = [temp for keys in temp_keys_by_density.values() for temp in keys]
    first_temp_keys = temp_keys_by_density[density_keys[0]]

    range_data = {
        'density_min': density_keys[0],
        'density_max': density_keys[-1],
        'temp_min': min(all_temps),
        'temp_max': max(all_temps),
        'density_step': round(density_keys[1] - density_keys[0], 4) if len(density_keys) > 1 else None,
        'temp_step': round(first_temp_keys[1] - first_temp_keys[0], 2) if len(first_temp_keys) > 1 else None,
    }
    return density_keys, temp_keys_by_density, range_data


def _parse_sheet(ws):
    temp_row = next(ws.iter_rows(min_row=9, max_row=9, values_only=True))
    temps = [float(value) for value in temp_row[3:] if value is not None]

    table = {}
    for row in ws.iter_rows(min_row=10, values_only=True):
        density_key = row[2]
        if density_key is None or not isinstance(density_key, (int, float)):
            continue

        density_key = round(float(density_key), 4)
        row_dict = {}
        for index, value in enumerate(row[3:]):
            if index < len(temps) and value is not None:
                row_dict[float(temps[index])] = float(value)
        if row_dict:
            table[density_key] = row_dict
    return table


def _interp1d(x0, x1, y0, y1, x):
    if x1 == x0:
        return y0
    return y0 + (y1 - y0) * (x - x0) / (x1 - x0)


def _lookup(table, density, temperature, density_keys, temp_keys_by_density, range_data):
    if not table or density is None or temperature is None or not range_data:
        return None

    if not (
        range_data['density_min'] <= density <= range_data['density_max']
        and range_data['temp_min'] <= temperature <= range_data['temp_max']
    ):
        return None

    density_index = bisect.bisect_left(density_keys, density)
    if density_index == 0:
        density_lo = density_hi = density_keys[0]
    elif density_index >= len(density_keys):
        density_lo = density_hi = density_keys[-1]
    elif density_keys[density_index] == density:
        density_lo = density_hi = density_keys[density_index]
    else:
        density_lo = density_keys[density_index - 1]
        density_hi = density_keys[density_index]

    def row_lookup(density_key, temp):
        row = table[density_key]
        temp_keys = temp_keys_by_density[density_key]
        temp_index = bisect.bisect_left(temp_keys, temp)

        if temp_index == 0:
            return row[temp_keys[0]]
        if temp_index >= len(temp_keys):
            return row[temp_keys[-1]]
        if temp_keys[temp_index] == temp:
            return row[temp_keys[temp_index]]

        temp_lo = temp_keys[temp_index - 1]
        temp_hi = temp_keys[temp_index]
        return _interp1d(temp_lo, temp_hi, row[temp_lo], row[temp_hi], temp)

    value_lo = row_lookup(density_lo, temperature)
    if density_lo == density_hi:
        return round(value_lo, 6)

    value_hi = row_lookup(density_hi, temperature)
    return round(_interp1d(density_lo, density_hi, value_lo, value_hi, density), 6)


def density_at_20_from_table(sample_density_kg_l, sample_temperature_c):
    _load_tables()
    if sample_density_kg_l is None or sample_temperature_c is None:
        return None

    result = _lookup(
        _TABLE_59B,
        sample_density_kg_l,
        sample_temperature_c,
        _TABLE_59B_KEYS,
        _TABLE_59B_TEMP_KEYS,
        _TABLE_59B_RANGE,
    )
    return round(result, 4) if result is not None else None


def vcf_from_table(density_at_20_kg_l, tank_temperature_c):
    _load_tables()
    if density_at_20_kg_l is None or tank_temperature_c is None:
        return None

    result = _lookup(
        _TABLE_60B,
        density_at_20_kg_l,
        tank_temperature_c,
        _TABLE_60B_KEYS,
        _TABLE_60B_TEMP_KEYS,
        _TABLE_60B_RANGE,
    )
    return round(result, 6) if result is not None else None


def wcf_from_density(density_at_20_kg_l):
    if density_at_20_kg_l is None:
        return None
    return round(max(density_at_20_kg_l - 0.0011, 0), 4)


def table_range():
    _load_tables()
    return _TABLE_59B_RANGE
