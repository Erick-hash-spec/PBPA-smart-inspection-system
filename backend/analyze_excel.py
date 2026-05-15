#!/usr/bin/env python
"""
Analyze the Shore Tank Calculation Excel file to understand structure
"""
import openpyxl
from openpyxl.styles import PatternFill

# Load the Excel file
wb = openpyxl.load_workbook('SHORE TANK CALCULATION EXCELL.xlsx')
ws = wb.active

print('Sheet name:', ws.title)
print('\nExcel File Analysis:')
print('=' * 140)
print(f'Max Row: {ws.max_row}, Max Column: {ws.max_column}')
print('=' * 140)

# Color mapping
color_map = {
    'FFFFFFFF': 'White (Label)',
    'FFFFFF00': 'Yellow (Manual Input)',
    '000070C0': 'Blue (Auto-Calc)',
    'FFD3D3D3': 'Grey (N/A)',
    'FF92D050': 'Green (Calculated)',
}

print('\n\nFIRST 35 ROWS:')
print('-' * 140)

for row_idx in range(1, min(36, ws.max_row + 1)):
    # Print row number
    row_content = f'Row {row_idx:2d}: '
    
    # Check columns
    for col_idx in range(1, min(10, ws.max_column + 1)):
        cell = ws.cell(row=row_idx, column=col_idx)
        value = cell.value if cell.value else '-'
        
        # Get fill color
        if cell.fill and cell.fill.start_color:
            color = cell.fill.start_color.rgb
            color_name = 'Unknown'
            for c, name in color_map.items():
                if str(color).upper().endswith(c.upper()) or c.upper() in str(color).upper():
                    color_name = name
                    break
        else:
            color_name = 'None'
        
        # Build display
        if value != '-':
            row_content += f'[{color_name}] {str(value)[:25]:25s} | '
        else:
            row_content += f'[{color_name}] {"-":25s} | '
    
    print(row_content[:140])

print('\n\nCOLOR LEGEND:')
print('-' * 140)
print('Yellow (FFFF00) = Manual Input (User enters values here)')
print('Blue   (0070C0) = Auto-Calculated (System calculates)')
print('Grey   (D3D3D3) = Not Applicable or Density @20°C & VCF (now auto-calculated via ASTM)')
print('White  (FFFFFF) = Labels/Headers')

# Find key sections
print('\n\nKEY SECTIONS FOUND:')
print('-' * 140)

section_keywords = ['SECTION', 'TABLE', 'TANK', 'DENSITY', 'VCF', 'WCF', 'TEMPERATURE', 'VOLUME', 'WEIGHT']
for row_idx in range(1, ws.max_row + 1):
    cell_value = str(ws.cell(row=row_idx, column=1).value or '')
    if any(keyword in cell_value.upper() for keyword in section_keywords):
        print(f'Row {row_idx}: {cell_value}')

print('\n✅ Analysis complete!\n')
